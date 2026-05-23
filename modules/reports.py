"""Модуль генерации отчётов — .xlsx и .pdf с поддержкой кириллицы."""
import os
import sys
import platform
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Flowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))
from logger import logger
from modules.analytics import get_student_cols, get_report_dir, today_str, is_debtor, grade_to_score


# ─── ШРИФТЫ С ПОДДЕРЖКОЙ КИРИЛЛИЦЫ ──────────────────────────────
# Регистрируем TTF-шрифты Windows / Linux / macOS

CYR_FONT = "Helvetica"        # имя шрифта в стилях ReportLab (по умолчанию)
CYR_FONT_BOLD = "Helvetica-Bold"
_FONTS_REGISTERED = False


def _register_cyrillic_fonts():
    """Регистрирует кириллический TTF-шрифт, доступный в системе."""
    global CYR_FONT, CYR_FONT_BOLD, _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return

    # Кандидаты для каждой ОС: (имя, путь к Regular, путь к Bold)
    if platform.system() == "Windows":
        candidates = [
            ("DejaVuSans",
             r"C:\Windows\Fonts\DejaVuSans.ttf",
             r"C:\Windows\Fonts\DejaVuSans-Bold.ttf"),
            ("Arial",
             r"C:\Windows\Fonts\arial.ttf",
             r"C:\Windows\Fonts\arialbd.ttf"),
            ("Tahoma",
             r"C:\Windows\Fonts\tahoma.ttf",
             r"C:\Windows\Fonts\tahomabd.ttf"),
            ("TimesNewRoman",
             r"C:\Windows\Fonts\times.ttf",
             r"C:\Windows\Fonts\timesbd.ttf"),
            ("Verdana",
             r"C:\Windows\Fonts\verdana.ttf",
             r"C:\Windows\Fonts\verdanab.ttf"),
        ]
    elif platform.system() == "Darwin":  # macOS
        candidates = [
            ("Arial",
             "/Library/Fonts/Arial.ttf",
             "/Library/Fonts/Arial Bold.ttf"),
            ("HelveticaNeue",
             "/System/Library/Fonts/Helvetica.ttc",
             "/System/Library/Fonts/Helvetica.ttc"),
        ]
    else:  # Linux
        candidates = [
            ("DejaVuSans",
             "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
             "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            ("LiberationSans",
             "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
             "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
        ]

    for name, regular, bold in candidates:
        if not os.path.isfile(regular):
            continue
        try:
            pdfmetrics.registerFont(TTFont(name, regular))
            bold_name = f"{name}-Bold"
            if os.path.isfile(bold):
                pdfmetrics.registerFont(TTFont(bold_name, bold))
            else:
                # Если нет жирной версии — используем обычную
                pdfmetrics.registerFont(TTFont(bold_name, regular))
            CYR_FONT = name
            CYR_FONT_BOLD = bold_name
            _FONTS_REGISTERED = True
            logger.info(f"PDF: зарегистрирован кириллический шрифт {name} ({regular})")
            return
        except Exception as e:
            logger.warning(f"PDF: шрифт {name} не зарегистрирован: {e}")
            continue

    logger.warning(
        "PDF: ни один кириллический TTF-шрифт не найден. "
        "Используется Helvetica (русский текст может отображаться как квадраты)."
    )


# Регистрируем шрифты при импорте модуля
_register_cyrillic_fonts()


# ─── EXCEL ─────────────────────────────────────────────────────
RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
HEAD = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ALT = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
HFONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _save_excel(df, path, highlight=True):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Успеваемость")
        ws = w.sheets["Успеваемость"]
        for ri in range(1, ws.max_row + 1):
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(ri, ci)
                cell.alignment = WRAP
                cell.border = THIN
                if ri == 1:
                    cell.fill = HEAD
                    cell.font = HFONT
                elif highlight and ci > 3 and is_debtor(str(cell.value or "")):
                    cell.fill = RED
                elif ri % 2 == 0:
                    cell.fill = ALT
        for ci, col in enumerate(df.columns, 1):
            width = min(max(df[col].astype(str).map(len).max(), len(str(col))) + 2, 45)
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.freeze_panes = "A2"


# ─── PDF ───────────────────────────────────────────────────────

class VerticalText(Flowable):
    """Вертикально повёрнутый текст (для шапки с ФИО студентов)."""
    def __init__(self, text, font_name=CYR_FONT_BOLD, font_size=7,
                 color=colors.white, height_cm=2.8):
        Flowable.__init__(self)
        self.text = text
        self.font_name = font_name
        self.font_size = font_size
        self.color = color
        # Высота ячейки шапки — в неё должны влезть длинные ФИО
        self.height = height_cm * cm

    def wrap(self, availWidth, availHeight):
        return (self.font_size * 1.2, self.height)

    def draw(self):
        canv = self.canv
        canv.saveState()
        canv.setFont(self.font_name, self.font_size)
        canv.setFillColor(self.color)
        # Поворот на 90° против часовой
        canv.rotate(90)
        # После rotate координаты переворачиваются:
        # x идёт вверх, y идёт влево
        canv.drawString(2, -self.font_size * 0.7, self.text)
        canv.restoreState()


def _save_pdf(df, path, title, group_name):
    doc = SimpleDocTemplate(
        path, pagesize=landscape(A4),
        leftMargin=1.0 * cm, rightMargin=1.0 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm
    )

    # Стили заголовков с кириллическим шрифтом
    ts = ParagraphStyle(
        "T", fontName=CYR_FONT_BOLD, fontSize=13,
        textColor=colors.HexColor("#1F4E79"), spaceAfter=4,
    )
    ss = ParagraphStyle(
        "S", fontName=CYR_FONT, fontSize=10,
        textColor=colors.HexColor("#666666"), spaceAfter=10,
    )

    # Стиль для текста в обычных ячейках (с переносом по словам)
    cell_style = ParagraphStyle(
        "Cell", fontName=CYR_FONT, fontSize=7,
        alignment=1,  # центр
        leading=8, wordWrap='CJK',
    )
    # Стиль для текстовых ячеек первых столбцов — выравнивание влево
    text_cell_style = ParagraphStyle(
        "TextCell", fontName=CYR_FONT, fontSize=7,
        alignment=0,  # слева
        leading=8, wordWrap='CJK',
    )
    # Стиль для шапки текстовых столбцов (Дисциплина, Тип, Семестр) — обычный
    header_text_style = ParagraphStyle(
        "HeaderText", fontName=CYR_FONT_BOLD, fontSize=8,
        textColor=colors.white, alignment=1,
        leading=10, wordWrap='CJK',
    )

    elems = [
        Paragraph(title, ts),
        Paragraph(
            f"Группа: {group_name} | {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            ss
        ),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1F4E79")),
        Spacer(1, 0.3 * cm),
    ]

    headers = list(df.columns)
    n_cols = len(headers)

    # Определяем индексы текстовых столбцов (Дисциплина, Тип, Семестр, Студент)
    TEXT_COL_NAMES = {"Дисциплина", "Тип", "Семестр", "Студент",
                      "Тип аттестации", "Оценка"}
    text_col_indices = {i for i, h in enumerate(headers) if h in TEXT_COL_NAMES}

    # ── Подготовка данных: оборачиваем каждую ячейку в Paragraph ──
    # Шапка: для текстовых столбцов — обычный Paragraph,
    # для столбцов со студентами — VerticalText (повёрнутый на 90°)
    header_row = []
    for i, h in enumerate(headers):
        if i in text_col_indices:
            header_row.append(Paragraph(str(h), header_text_style))
        else:
            # Это столбец с ФИО студента — рисуем вертикально
            header_row.append(VerticalText(str(h)))

    body_rows = []
    for r in df.itertuples(index=False):
        row_cells = []
        for ci, val in enumerate(r):
            text = str(val) if val is not None else ""
            if ci in text_col_indices:
                row_cells.append(Paragraph(text, text_cell_style))
            else:
                row_cells.append(Paragraph(text, cell_style))
        body_rows.append(row_cells)

    data = [header_row] + body_rows

    # ── Расчёт ширин столбцов ──
    pw = landscape(A4)[0] - 2 * cm  # доступная ширина страницы

    # Сколько текстовых столбцов в начале (Дисциплина/Тип/Семестр/...)
    # — им даём фиксированную ширину
    n_text_cols = sum(1 for i in range(min(3, n_cols)) if i in text_col_indices)

    if n_text_cols >= 3:
        # Сводная ведомость: Дисциплина(4.5см) + Тип(2см) + Семестр(1.8см)
        fixed_widths = [4.5 * cm, 2.0 * cm, 1.8 * cm]
    elif n_text_cols == 2:
        fixed_widths = [4.5 * cm, 2.0 * cm]
    elif n_text_cols == 1:
        fixed_widths = [4.5 * cm]
    else:
        fixed_widths = []

    remaining_cols = n_cols - len(fixed_widths)
    remaining_width = pw - sum(fixed_widths)
    student_col_width = remaining_width / remaining_cols if remaining_cols > 0 else 1 * cm

    # Если столбцов очень много (>15 студентов) — уменьшим минимум, но не меньше 0.8см
    student_col_width = max(student_col_width, 0.8 * cm)

    col_widths = fixed_widths + [student_col_width] * remaining_cols

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    st = TableStyle([
        # Заголовок
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
        # Данные
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 1), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
    ])

    # Подсветка задолженностей
    for ri, _ in enumerate(df.itertuples(), 1):
        for ci, col in enumerate(df.columns):
            if ci >= 3 and is_debtor(str(df.iloc[ri - 1, ci])):
                st.add("BACKGROUND", (ci, ri), (ci, ri),
                       colors.HexColor("#FFCCCC"))

    tbl.setStyle(st)
    elems.append(tbl)
    doc.build(elems)


# ─── ПУБЛИЧНЫЕ ФУНКЦИИ ─────────────────────────────────────────
def report_debtors(df, group_name, num_sem=None):
    import re
    student_cols = get_student_cols(df)
    if num_sem is None:
        nums = [int(m.group(1)) for s in df["Семестр"].unique()
                if (m := re.search(r"(\d+)", str(s)))]
        num_sem = max(nums) if nums else 1
    sem_df = df[df["Семестр"].str.contains(f"{num_sem} семестр", na=False)]
    rows = []
    for _, row in sem_df.iterrows():
        for s in student_cols:
            if is_debtor(row[s]):
                rows.append({
                    "Студент": s,
                    "Дисциплина": row["Дисциплина"],
                    "Тип аттестации": row["Тип"],
                    "Семестр": row["Семестр"],
                    "Оценка": row[s],
                })
    if not rows:
        logger.info(f"Должников за {num_sem} семестр не найдено")
        return []
    d = pd.DataFrame(rows)
    rd = get_report_dir(group_name, "Должники")
    dt = today_str()
    xp = os.path.join(rd, f"должники_{num_sem}сем_{group_name}_{dt}.xlsx")
    pp = os.path.join(rd, f"должники_{num_sem}сем_{group_name}_{dt}.pdf")
    _save_excel(d, xp, highlight=False)
    _save_pdf(d, pp, f"Отчёт по должникам — {num_sem} семестр", group_name)
    logger.info(f"Отчёт по должникам: {len(rows)} записей → {xp}")
    return [xp, pp]


def report_full_period(df, group_name, individual=False):
    student_cols = get_student_cols(df)
    rd = get_report_dir(group_name, "Весь_период")
    dt = today_str()
    files = []
    if individual:
        for s in student_cols:
            sd = df[["Семестр", "Дисциплина", "Тип", s]].rename(columns={s: "Оценка"})
            sn = s.replace(" ", "_")
            xp = os.path.join(rd, f"{sn}_период_{dt}.xlsx")
            pp = os.path.join(rd, f"{sn}_период_{dt}.pdf")
            _save_excel(sd, xp)
            _save_pdf(sd, pp, f"Успеваемость: {s}", group_name)
            files += [xp, pp]
    else:
        dc = df.copy()
        ar = {"Дисциплина": "Средний балл", "Тип": "", "Семестр": ""}
        for s in student_cols:
            ar[s] = round(df[s].apply(grade_to_score).mean(), 2)
        dc = pd.concat([dc, pd.DataFrame([ar])], ignore_index=True)
        xp = os.path.join(rd, f"весь_период_{group_name}_{dt}.xlsx")
        pp = os.path.join(rd, f"весь_период_{group_name}_{dt}.pdf")
        _save_excel(dc, xp)
        _save_pdf(dc, pp, "Сводная ведомость успеваемости", group_name)
        files = [xp, pp]
    logger.info(f"Отчёт за весь период: {len(files)} файлов")
    return files


def report_by_semester(df, group_name, semester_number, individual=False):
    sd = df[df["Семестр"].str.contains(f"{semester_number} семестр", na=False)].copy()
    rd = get_report_dir(group_name, "По_семестрам")
    dt = today_str()
    files = []
    if individual:
        for s in get_student_cols(df):
            fd = sd[["Дисциплина", "Тип", s]].rename(columns={s: "Оценка"})
            xp = os.path.join(rd, f"{s.replace(' ', '_')}_{semester_number}сем_{dt}.xlsx")
            _save_excel(fd, xp)
            files.append(xp)
    else:
        xp = os.path.join(rd, f"отчёт_{semester_number}сем_{group_name}_{dt}.xlsx")
        pp = os.path.join(rd, f"отчёт_{semester_number}сем_{group_name}_{dt}.pdf")
        _save_excel(sd, xp)
        _save_pdf(sd, pp, f"Успеваемость за {semester_number} семестр", group_name)
        files = [xp, pp]
    logger.info(f"Отчёт за {semester_number} семестр: {files}")
    return files

def report_student(df, group_name: str, student_name: str, num_sem=None):
    """Отчёт по конкретному студенту — поиск по подстроке ФИО."""
    import re

    student_cols = get_student_cols(df)

    # Ищем столбец (регистронезависимо, по подстроке)
    matched = [c for c in student_cols if student_name.lower() in c.lower()]

    if not matched:
        # Пробуем частичное совпадение по первым буквам слова
        matched = [
            c for c in student_cols
            if any(student_name.lower() in w.lower() for w in c.split())
        ]

    if not matched:
        logger.warning(f"Студент «{student_name}» не найден. Доступны: {student_cols[:5]}")
        return [], f"Студент «{student_name}» не найден в группе {group_name}."

    col = matched[0]  # берём первое совпадение

    # Фильтруем по семестру если задан
    work_df = df.copy()
    if num_sem:
        work_df = work_df[work_df["Семестр"].str.contains(f"{num_sem} семестр", na=False)]
        if work_df.empty:
            return [], f"Нет данных за {num_sem} семестр для студента {col}."

    # Формируем таблицу: Семестр / Дисциплина / Тип / Оценка
    student_df = work_df[["Семестр", "Дисциплина", "Тип", col]].rename(columns={col: "Оценка"})

    # Имя файла — безопасное (убираем спецсимволы)
    safe_name = re.sub(r"[^\w]", "_", col)
    sem_label = f"_{num_sem}сем" if num_sem else "_все_сем"
    rd = get_report_dir(group_name, "По_студентам")
    dt = today_str()

    xp = os.path.join(rd, f"{safe_name}{sem_label}_{dt}.xlsx")
    pp = os.path.join(rd, f"{safe_name}{sem_label}_{dt}.pdf")

    title = f"Успеваемость: {col}"
    if num_sem:
        title += f" · {num_sem} семестр"
    else:
        title += " · весь период обучения"

    _save_excel(student_df, xp, highlight=False)
    _save_pdf(student_df, pp, title, group_name)

    logger.info(f"Отчёт по студенту {col}: {len(student_df)} строк → {xp}")
    return [xp, pp], col