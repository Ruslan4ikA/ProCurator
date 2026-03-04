"""
Модуль генерации отчетов в форматах Excel и PDF
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from utils.helpers import (
    is_debtor, get_student_columns, sanitize_filename, get_current_date
)
from config import REPORTS_DIR

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Генератор отчетов"""
    
    def __init__(self):
        self.red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        self.green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    
    def _format_worksheet(self, worksheet, highlight_debtors: bool = True, 
                         individual: bool = False, student_name: str = None):
        """Форматирование листа Excel"""
        wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Определяем столбцы студентов
        student_cols = get_student_columns(pd.read_excel(worksheet.parent.filename))
        
        # Проходим по строкам
        for row_idx in range(2, worksheet.max_row + 1):
            # Определяем, есть ли долг в строке
            has_debt_in_row = False
            for col_idx in range(4, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value and is_debtor(cell_value):
                    has_debt_in_row = True
                    break
            
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = wrap_align
                
                if highlight_debtors:
                    if individual and has_debt_in_row:
                        # Подсвечиваем ВСЮ строку для индивидуальных отчетов
                        cell.fill = self.red_fill
                    elif not individual and col_idx >= 4:
                        # Подсвечиваем ТОЛЬКО ячейку с долгом
                        if cell.value and is_debtor(cell.value):
                            cell.fill = self.red_fill
        
        # Форматирование заголовка
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # Автоподбор ширины столбцов
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 0
            column = worksheet.column_dimensions[get_column_letter(col_idx)]
            for row_idx in range(1, worksheet.max_row + 1):
                cell_value = str(worksheet.cell(row=row_idx, column=col_idx).value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            
            adjusted_width = min(max_length + 2, 50)
            column.width = adjusted_width
    
    def generate_group_report(self, df: pd.DataFrame, group_name: str, 
                             semester: str = None) -> Path:
        """Генерация общего отчета по группе"""
        # Добавляем средний балл
        from modules.analyzer import GradeAnalyzer
        df_with_avg = GradeAnalyzer.calculate_average_score(df)
        
        # Формируем имя файла
        date_str = get_current_date()
        filename = f"весь_период_группа_{group_name}_{date_str}.xlsx"
        if semester:
            filename = f"отчет_{semester}_семестр_{group_name}_{date_str}.xlsx"
        
        filepath = REPORTS_DIR / group_name / filename
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Сохраняем в Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_with_avg.to_excel(writer, index=False, sheet_name='Отчет')
            worksheet = writer.sheets['Отчет']
            self._format_worksheet(worksheet, highlight_debtors=True, individual=False)
        
        logger.info(f"Общий отчет сохранен: {filepath}")
        return filepath
    
    def generate_individual_reports(self, df: pd.DataFrame, group_name: str,
                                   semester: str = None) -> list:
        """Генерация индивидуальных отчетов по каждому студенту"""
        student_cols = get_student_columns(df)
        generated_files = []
        
        date_str = get_current_date()
        
        for student in student_cols:
            # Создаем индивидуальный отчет
            student_data = df[['Семестр', 'Дисциплина', 'Тип', student]].copy()
            student_data = student_data.rename(columns={student: 'Оценка'})
            
            # Формируем имя файла
            safe_student_name = sanitize_filename(student)
            filename = f"{safe_student_name}_весь_период_{date_str}.xlsx"
            if semester:
                filename = f"{safe_student_name}_{semester}_семестр_{date_str}.xlsx"
            
            filepath = REPORTS_DIR / group_name / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # Сохраняем в Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                student_data.to_excel(writer, index=False, sheet_name='Отчет')
                worksheet = writer.sheets['Отчет']
                self._format_worksheet(worksheet, highlight_debtors=True, individual=True)
            
            generated_files.append(filepath)
        
        logger.info(f"Создано {len(generated_files)} индивидуальных отчетов")
        return generated_files
    
    def generate_debtors_report(self, debt_list, group_name: str,
                               semester: str = None) -> Path:
        """Генерация отчета по должникам"""
        if not debt_list:
            logger.warning("Нет данных о должниках для генерации отчета")
            return None
        
        debt_df = pd.DataFrame(debt_list)
        
        # Формируем имя файла
        date_str = get_current_date()
        filename = f"должники_{group_name}_{date_str}.xlsx"
        if semester:
            filename = f"должники_{semester}_семестр_{group_name}_{date_str}.xlsx"
        
        filepath = REPORTS_DIR / group_name / filename
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Сохраняем в Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            debt_df.to_excel(writer, index=False, sheet_name='Должники')
            worksheet = writer.sheets['Должники']
            self._format_worksheet(worksheet, highlight_debtors=False)
        
        logger.info(f"Отчет о должниках сохранен: {filepath}")
        return filepath